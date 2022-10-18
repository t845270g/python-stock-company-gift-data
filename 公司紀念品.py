from openpyxl.utils import get_column_letter
from numpy import append
import requests #匯入爬取模組
from openpyxl import Workbook,load_workbook
wb=Workbook()#創建新表格檔案
ws=wb.active#叫出當前執行工作區域
ws.title="股東會紀念品"


res = requests.get('https://histock.tw/stock/gift.aspx')#get("網址")爬取指定網址資料存在res記憶體中
res.headers#取得回應的headers
res.encoding = 'utf-8'#重新解碼，使用utf-8進行解碼才能顯示中文字

from bs4 import  BeautifulSoup
soup=BeautifulSoup(res.text,'html.parser') #運用html.parser解析原始碼

lm=soup.find("table",{"id":"CPHB1_gvOld"})#尋找指定標籤與id名稱，得到買進日到期的所有資料


lm=soup.select(".gvTB")[1]#尋找指定標籤與id名稱，得到買進日到期的所有資料
lm2=lm.find_all('tr')#[0]是表格的樣式與屬性名稱，不取，從[1]開始為所有資料
編號=['A','B','C','D','E','F','G','H','I','J','K','L','M','N']




for x in range(1,len(lm2)):
    if bool(lm2[x].find_all('td')[0])==True:
        for y in range(12):   
            if y == 0:
                ws[編號[y]+str(x)]=int(lm2[x].find_all('td')[y].text)
            else:
                ws[編號[y]+str(x)]=lm2[x].find_all('td')[y].text
    else:
        break    





wb.save(r"股東會紀念品資料.xlsx")#儲存新表格